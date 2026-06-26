# -*- coding: utf-8 -*-
"""工资表 → 社保测算标准模板 · 核心逻辑（与 Claude 技能同口径，可跨工具复用）。

标准模板 10 列：
  工号/唯一标识 | 姓名 | 所属主体 | 所属年月(YYYY-MM) | 应发工资(税前合计) |
  身份类型 | 实发工资(可选) | 是否缴公积金 | 公积金缴存基数 | 公积金缴存比例
"""
import os, re, csv, glob
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SYN = {
    'name':   ['姓名', '员工姓名', '名字', '人员姓名', '职工姓名', 'name', '员工'],
    'id':     ['身份证', '身份证号', '身份证号码', '工号', '员工编号', '员工工号', '人员编号',
               '帐号', '账号', '工号/唯一标识', '唯一标识', '手机号', '手机号码', 'id', 'empid'],
    'entity': ['公司', '所属主体', '主体', '所属公司', '单位名称', '所属单位', '分公司', '法人主体'],
    'ym':     ['所属年月', '年月', '月份', '工资月份', '所属月份', '发放月份', '薪资月份', '会计期间'],
    'gross':  ['应发工资', '应发合计', '应发', '应发工资(税前合计)', '应发工资（税前合计）',
               '税前工资', '税前合计', '应发金额', '应付工资'],
    'total':  ['工资合计', '合计', '工资总额', '总计'],
    'unit_ss':['社保单位缴纳', '单位社保', '社保(单位)', '单位缴纳社保', '社保单位', '单位承担社保', '公司社保'],
    'net':    ['实发金额', '实发工资', '实发', '实发合计', '到手', '实发数', '净发'],
    'stype':  ['身份类型', '用工类型', '人员类型', '员工类型', '用工性质'],
    'status': ['状态', '在职状态', '员工状态', '用工状态', '人员状态'],
    'gjj_paid':['是否缴公积金', '公积金缴存', '是否缴存公积金', '是否缴纳公积金'],
    'gjj_base':['公积金缴存基数', '公积金基数', '住房公积金基数'],
    'gjj_ratio':['公积金缴存比例', '公积金比例', '住房公积金比例'],
}
EARN_KW   = ['基本工资', '基础工资', '岗位工资', '岗位津贴', '绩效', '提成', '奖金', '奖励',
            '津贴', '补贴', '加班', '工龄', '全勤', '话费', '车补', '餐补', '高温', '其他奖补']
DEDUCT_KW = ['迟到', '缺卡', '缺勤', '请假', '旷工', '事假', '病假', '假勤', '扣款', '扣发', '罚款']
EXCLUDE_KW= ['社保', '公积金', '个税', '个人所得税', '实发', '应发', '合计', '基数', '比例', '账号', '帐号', '工号', '银行']
STYPES = ['全日制正式', '试用期', '老板/股东', '退休返聘', '在校实习生', '已参保兼职', '劳务外包', '合规非全日制']
# 前3类为应参保身份；其余5类不纳入参保测算。用于「应参保却无单位社保」交叉校验。
PARTICIPATING = {'全日制正式', '试用期', '老板/股东'}
# 工资表「状态」列 → 标准身份类型。按关键字子串匹配（"试用期离职"也能命中）。
# 注意：「离职」是用工生命周期、非身份类型，无法据此判定参保身份，故不映射，保持默认并在报告中提示。
STATUS_MAP = [
    ('外包', '劳务外包'),
    ('劳务', '劳务外包'),
    ('实习', '在校实习生'),
    ('退休', '退休返聘'),
    ('返聘', '退休返聘'),
    ('兼职', '已参保兼职'),
    ('股东', '老板/股东'),
    ('老板', '老板/股东'),
    ('非全日制', '合规非全日制'),
    ('试用', '试用期'),
    ('正式', '全日制正式'),
    ('在职', '全日制正式'),
]


def map_status(raw, extra=None):
    """把原表「状态」值映射为标准身份类型。
    extra: 可选 [(关键字, 身份类型)] 或 {关键字: 身份类型}，按公司补充的自定义词表，优先于内置 STATUS_MAP。
    命中返回(身份类型, 原值)，未命中返回(None, 原值)。"""
    if raw is None:
        return None, ''
    s = str(raw).strip()
    if not s:
        return None, ''
    pairs = list(extra.items()) if isinstance(extra, dict) else (list(extra) if extra else [])
    for kw, st in pairs + STATUS_MAP:
        if kw and kw in s and st in STYPES:
            return st, s
    return None, s
COLS = ['工号/唯一标识', '姓名', '所属主体', '所属年月', '应发工资(税前合计)',
        '身份类型', '实发工资(可选)', '是否缴公积金', '公积金缴存基数', '公积金缴存比例']
FIELD_LABELS = {'id': '工号/唯一标识', 'name': '姓名', 'entity': '所属主体', 'ym': '所属年月',
                'gross': '应发工资', 'total': '工资合计', 'unit_ss': '单位社保', 'net': '实发',
                'stype': '身份类型', 'gjj_paid': '是否缴公积金', 'gjj_base': '公积金基数', 'gjj_ratio': '公积金比例'}


def norm(s):
    return re.sub(r'[\s　()（）\-_/]', '', str(s)).lower() if s is not None else ''


def build_index(header):
    idx = {}
    nh = [norm(h) for h in header]
    for key, names in SYN.items():
        for nm in names:
            n = norm(nm)
            for j, h in enumerate(nh):
                if h == n and key not in idx:
                    idx[key] = j
    return idx


def earn_deduct_cols(header):
    nh = [norm(h) for h in header]
    earn, ded = [], []
    for j, h in enumerate(nh):
        if not h:
            continue
        if any(k in h for k in EXCLUDE_KW):
            continue
        if any(norm(k) in h for k in DEDUCT_KW):
            ded.append(j)
        elif any(norm(k) in h for k in EARN_KW):
            earn.append(j)
    return earn, ded


def read_rows(path):
    """返回 (header:list, rows:list[list], err:str|None)"""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == '.csv':
            with open(path, encoding='utf-8-sig') as f:
                data = [r for r in csv.reader(f)]
        elif ext == '.xls':
            try:
                import xlrd
            except ImportError:
                return None, [], "旧版 .xls 需要 xlrd(pip install xlrd) 或先另存为 .xlsx"
            book = xlrd.open_workbook(path); sh = book.sheet_by_index(0)
            data = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        else:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            data = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    except Exception as e:
        return None, [], f"读取失败: {e}"
    name_set = {norm(x) for x in SYN['name']}
    hr = None
    for i, r in enumerate(data):
        if r and any(norm(c) in name_set for c in r):
            hr = i; break
    if hr is None:
        return None, [], "未找到含「姓名」的表头行"
    return data[hr], data[hr + 1:], None


def infer_ym(fname, idx, row):
    if 'ym' in idx and idx['ym'] < len(row) and row[idx['ym']]:
        v = str(row[idx['ym']])
        m = re.search(r'(\d{4})\D?(\d{1,2})', v)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月', fname)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    return ''


def infer_entity(fname, idx, row):
    if 'entity' in idx and idx['entity'] < len(row) and row[idx['entity']]:
        return str(row[idx['entity']]).strip()
    base = os.path.splitext(os.path.basename(fname))[0]
    seg = base.split('_')[-1] if '_' in base else base
    seg = re.sub(r'\d{4}\s*年\s*\d{1,2}\s*月', '', seg)
    seg = re.sub(r'\d{4}[-./]\d{1,2}', '', seg)
    seg = re.sub(r'(工资表|工资|薪资表|薪资|薪酬表|薪酬|含个税|明细表|明细|报表|表)', '', seg).strip(' -_')
    return seg or base


def num(v):
    if v is None or v == '':
        return None
    try:
        return float(str(v).replace(',', '').replace('¥', '').strip())
    except Exception:
        return None


def _gross_method(idx, earn_cols):
    if 'gross' in idx:
        return '直接取应发列'
    if 'total' in idx and 'unit_ss' in idx:
        return '工资合计−单位社保'
    if 'total' in idx:
        return '工资合计(无单位社保列,按原值)'
    if earn_cols:
        return '组件求和(收入项合计−考勤扣款)·估算待核'
    return '缺应发列(空)'


def list_files(folder):
    return sorted([f for f in glob.glob(os.path.join(folder, '*'))
                   if os.path.splitext(f)[1].lower() in ('.xlsx', '.xls', '.csv')
                   and not os.path.basename(f).startswith('~$')
                   and '标准模板' not in os.path.basename(f) and '整理结果' not in os.path.basename(f)])


def inspect(folder):
    """逐个文件返回：表头、样本行、自动识别的字段映射、应发口径、问题。
    供 AI 客户端判断非标表头后用 normalize 的 overrides 修正。"""
    out = []
    for path in list_files(folder):
        fn = os.path.basename(path)
        header, rows, err = read_rows(path)
        if header is None:
            out.append({'file': fn, 'error': err, 'header': None, 'sample_rows': [], 'detected_mapping': {}})
            continue
        idx = build_index(header)
        earn_cols, _ = earn_deduct_cols(header)
        mapping = {FIELD_LABELS.get(k, k): (header[v] if v < len(header) else None) for k, v in idx.items()}
        samples = [[('' if c is None else c) for c in r] for r in rows[:3]]
        out.append({
            'file': fn,
            'header': [('' if h is None else h) for h in header],
            'sample_rows': samples,
            'detected_mapping': mapping,
            'gross_method': _gross_method(idx, earn_cols),
            'has_name_col': 'name' in idx,
            'has_id_col': 'id' in idx,
            'has_status_col': 'status' in idx,
            'status_values': sorted({str(r[idx['status']]).strip() for r in rows
                                     if 'status' in idx and idx['status'] < len(r)
                                     and r[idx['status']] not in (None, '')})[:15] if 'status' in idx else [],
        })
    return out


def _resolve_overrides(header, ov):
    """把 override 的 column_map(字段->表头串或列序号) 解析为 idx 增量。"""
    extra = {}
    nh = [norm(h) for h in header]
    for field, ref in (ov.get('column_map') or {}).items():
        if isinstance(ref, int):
            extra[field] = ref
        else:
            n = norm(ref)
            for j, h in enumerate(nh):
                if h == n:
                    extra[field] = j
                    break
    return extra


def normalize(folder, output_dir=None, overrides=None):
    """整理文件夹内所有工资表为标准模板 + 报告。
    overrides: { 文件名: { 'entity': str, 'ym': 'YYYY-MM', 'column_map': {字段: 表头串或列号},
                          'status_map': {状态原值关键字: 标准身份类型} } }
      另支持全局键 '__status_map__': {关键字: 身份类型}，对所有文件生效（各公司状态写法不一时补词表用）。
      身份类型须取自 STYPES；status_map 优先于内置关键字表。
    返回 dict（含产出路径、行数、口径分布、需人工项、未识别状态值、报告 markdown）。"""
    overrides = overrides or {}
    global_status_map = overrides.get('__status_map__', {}) or {}
    output_dir = output_dir or folder
    files = list_files(folder)
    if not files:
        raise FileNotFoundError(f"在 {folder} 没找到工资表(.xlsx/.xls/.csv)")

    records, report = [], []
    id_sources, no_id = set(), []
    gross_methods, empty_gross = defaultdict(int), {}
    need_manual = []
    stype_source = defaultdict(int)          # 身份类型来源：状态列推断 / 身份类型列 / 默认(无状态/离职待定/状态未识别)
    status_persons = defaultdict(set)        # 标准身份类型 -> {姓名} （由状态推断而来，供人工复核）
    left_persons = set()                     # 状态=离职，无法判定身份，需人工确认
    unknown_status = defaultdict(set)        # 未识别的状态原值 -> {姓名}（已暂按全日制正式，绝不静默，待补 status_map）
    suspect_rows = []                        # 有姓名但应发算不出来的行（疑似年终奖/汇总/合并单元格/错位），dump 原始单元格
    ss_gap = defaultdict(lambda: {'zero': [], 'paid': 0})  # (姓名,主体) -> 应参保身份却无单位社保的(年月)清单 + 有缴月数
    report.append(f"# 工资表整理报告\n\n输入: `{folder}`  共 {len(files)} 个文件\n")

    for path in files:
        fn = os.path.basename(path)
        ov = overrides.get(fn, {})
        smap = {**global_status_map, **(ov.get('status_map') or {})}   # 全局词表 + 本文件词表
        header, rows, err = read_rows(path)
        if header is None:
            report.append(f"- ⚠️ `{fn}`：{err}，**已跳过（建议 AI 读取此文件并用 overrides 映射）**。")
            need_manual.append((fn, err)); continue
        idx = build_index(header)
        idx.update(_resolve_overrides(header, ov))
        if 'name' not in idx:
            report.append(f"- ⚠️ `{fn}`：识别到表头但无「姓名」列，**已跳过（需 overrides 映射 name 列）**。")
            need_manual.append((fn, '无姓名列')); continue
        earn_cols, ded_cols = earn_deduct_cols(header)
        gmethod = _gross_method(idx, earn_cols)
        if 'id' in idx:
            id_sources.add(norm(header[idx['id']]))
        cnt, file_empty = 0, 0
        for r in rows:
            if not r or idx['name'] >= len(r) or not r[idx['name']] or str(r[idx['name']]).strip() in ('', '合计', '总计', '小计', '备注'):
                continue
            cell = lambda j: r[j] if j is not None and j < len(r) else None
            name = str(r[idx['name']]).strip()
            uid = str(cell(idx.get('id'))).strip() if cell(idx.get('id')) not in (None, '') else ''
            if not uid:
                uid = name
                if name not in no_id:
                    no_id.append(name)
            if gmethod == '直接取应发列':
                gross = num(cell(idx['gross']))
            elif gmethod == '工资合计−单位社保':
                t, u = num(cell(idx['total'])), num(cell(idx['unit_ss'])) or 0
                gross = (t - u) if t is not None else None
            elif gmethod.startswith('工资合计'):
                gross = num(cell(idx['total']))
            elif gmethod.startswith('组件求和'):
                e = sum(num(cell(j)) or 0 for j in earn_cols)
                d = sum(abs(num(cell(j)) or 0) for j in ded_cols)
                gross = e - d
            else:
                gross = None
            unit_ss = num(cell(idx.get('unit_ss')))   # 单位社保(缴纳)，用于「应参保却无社保」交叉校验
            if gross is None or gross == '':
                file_empty += 1
                # 有姓名却算不出应发：多为年终奖/汇总/合并单元格/列错位行，dump 原值供人工判断（不静默成空记录）
                raw = {(header[j] if j < len(header) and header[j] else f'col{j}'): v
                       for j, v in enumerate(r) if v not in (None, '')}
                suspect_rows.append({'file': fn, 'name': name, 'cells': raw})
                continue   # 不把算不出应发的行写进标准表，改由报告「疑似非工资行」暴露
            net = num(cell(idx.get('net')))
            # 身份类型优先级：① 原表「身份类型」列显式值 → ② 由「状态」列推断 → ③ 默认全日制正式
            stype = str(cell(idx.get('stype'))).strip() if cell(idx.get('stype')) else ''
            if stype in STYPES:
                stype_source['身份类型列'] += 1
            else:
                raw_status = cell(idx.get('status'))
                mapped, raw_s = map_status(raw_status, smap)
                if mapped:
                    stype = mapped
                    stype_source['状态列推断'] += 1
                    status_persons[mapped].add(name)
                else:
                    stype = '全日制正式'   # 兜底，但下面会分类标注，不静默
                    if not raw_s:
                        stype_source['默认(无状态列/空)'] += 1
                    elif '离职' in raw_s:
                        stype_source['默认(离职待定)'] += 1
                        left_persons.add(name)
                    else:
                        stype_source['默认(状态未识别)'] += 1
                        unknown_status[raw_s].add(name)
            gp = ''
            if cell(idx.get('gjj_paid')) not in (None, ''):
                gp = '是' if str(cell(idx.get('gjj_paid'))).strip() in ('是', 'Y', 'y', '1', 'TRUE', 'true', '缴') else '否'
            gb = num(cell(idx.get('gjj_base')))
            gr = num(cell(idx.get('gjj_ratio')))
            entity = ov.get('entity') or infer_entity(path, idx, r)
            ym = ov.get('ym') or infer_ym(path, idx, r)
            # 交叉校验：应参保身份(前3类)却没有单位社保 → 漏缴/身份分错的信号
            if stype in PARTICIPATING and (gross is not None):
                if unit_ss and unit_ss > 0:
                    ss_gap[(name, entity)]['paid'] += 1
                else:
                    ss_gap[(name, entity)]['zero'].append(ym)
            records.append([uid, name, entity, ym,
                            round(gross, 2) if gross is not None else '', stype,
                            round(net, 2) if net is not None else '',
                            gp, round(gb, 2) if gb is not None else '', gr if gr is not None else ''])
            cnt += 1
        gross_methods[gmethod] += 1
        if file_empty:
            empty_gross[fn] = file_empty
        if gmethod.startswith('组件求和') or file_empty:
            need_manual.append((fn, gmethod if gmethod.startswith('组件求和') else f'{file_empty}行空应发'))
        flag = ' ⚠️应发待核' if gmethod.startswith('组件求和') or file_empty else ''
        report.append(f"- ✅ `{fn}`：{cnt} 行 | 主体「{ov.get('entity') or infer_entity(path, idx, rows[0] if rows else [])}」 | 应发口径：{gmethod}{flag} | ID列：{header[idx['id']] if 'id' in idx else '无(用姓名兜底)'}")

    os.makedirs(output_dir, exist_ok=True)
    out_xlsx = os.path.join(output_dir, '社保测算标准模板_整理结果.xlsx')
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '标准模板'
    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        ws.cell(1, c).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, c).fill = PatternFill('solid', fgColor='3A5878')
    for rec in records:
        ws.append(rec)
    for i, w in enumerate([16, 9, 14, 11, 16, 12, 13, 12, 13, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    dv = DataValidation(type='list', formula1='"%s"' % ','.join(STYPES), allow_blank=False)
    ws.add_data_validation(dv); dv.add(f'F2:F{max(ws.max_row, 2)}')
    dvg = DataValidation(type='list', formula1='"是,否"', allow_blank=True)
    ws.add_data_validation(dvg); dvg.add(f'H2:H{max(ws.max_row, 2)}')
    ws.freeze_panes = 'A2'
    wb.save(out_xlsx)

    report.append(f"\n## 汇总\n- 总记录数：**{len(records)}** 行（人×主体×月）")
    report.append(f"- 唯一标识来源列：{', '.join(sorted(id_sources)) or '无'}")
    report.append(f"- 应发口径分布：{dict(gross_methods)}")
    report.append(f"- 身份类型来源：{dict(stype_source)}")
    report.append("\n## 🧭 身份类型识别（据原表「状态」列推断）")
    if status_persons:
        for st in sorted(status_persons):
            ppl = sorted(status_persons[st])
            report.append(f"- **{st}**（{len(ppl)}人）：{'、'.join(ppl)}")
    else:
        report.append("- 未识别到「状态」列，全部按默认「全日制正式」处理。")
    if left_persons:
        report.append(f"- ⚠️ **状态=离职，身份类型待人工确认**（{len(left_persons)}人，已暂按全日制正式）："
                      f"{'、'.join(sorted(left_persons))} —— 离职是用工生命周期、非身份类型，请按其在职期间实际身份(正式/试用)修正。")
    if unknown_status:
        report.append("\n## 🚨 未识别的「状态」值（已暂按全日制正式，**绝非静默**，请补 status_map）")
        report.append("各公司状态写法不一，下列值内置词表认不出。判断其对应身份类型后，重跑 `normalize_payroll` 并传 "
                      "`overrides`：全局用 `\"__status_map__\": {\"原值\": \"标准身份类型\"}`，或按文件用 `\"文件名\": {\"status_map\": {...}}`。")
        for val in sorted(unknown_status):
            ppl = sorted(unknown_status[val])
            report.append(f"- 状态「**{val}**」（{len(ppl)}人）：{'、'.join(ppl[:20])}{'…' if len(ppl) > 20 else ''}")
        report.append(f"  可选标准身份类型：{' / '.join(STYPES)}")
    # 应参保身份却无单位社保：never=从未缴(高风险,疑似身份分错或全程漏缴)；partial=部分月缺(入离职过渡/漏缴)
    never_ss = {k: v for k, v in ss_gap.items() if v['zero'] and v['paid'] == 0}
    partial_ss = {k: v for k, v in ss_gap.items() if v['zero'] and v['paid'] > 0}
    if never_ss or partial_ss:
        report.append("\n## 🔔 应参保身份却无单位社保（漏缴 / 身份分错 信号）")
        report.append("身份属前3类(全日制正式/试用期/老板股东)本应参保，但原表「单位社保」为 0/空。"
                      "请核实：是身份分错了(如其实是实习生/外包/非全日制)，还是确有漏缴。")
        if never_ss:
            report.append(f"- 🚨 **从未缴单位社保**（{len(never_ss)}人·主体，高风险）：")
            for (nm, ent), v in sorted(never_ss.items()):
                report.append(f"    - {nm}（{ent}）：{len(v['zero'])}个月全无社保 —— {('、'.join(v['zero']))}")
        if partial_ss:
            report.append(f"- ⚠️ **部分月份缺单位社保**（{len(partial_ss)}人·主体，多为入/离职过渡，仍请抽查）：")
            for (nm, ent), v in sorted(partial_ss.items()):
                report.append(f"    - {nm}（{ent}）：缺{len(v['zero'])}月 / 有{v['paid']}月，缺的月份：{('、'.join(v['zero']))}")
    if suspect_rows:
        report.append("\n## 🧩 疑似非工资行 / 列错位行（年终奖、汇总、合并单元格等）")
        report.append("下列行有姓名但应发算不出来，已**未计入**标准表，原始单元格如下，请人工判断如何处理"
                      "（并入当月应发 / 单列 / 剔除），必要时改源表或用 overrides 重跑：")
        for s in suspect_rows:
            kv = '，'.join(f"{k}={v}" for k, v in s['cells'].items())
            report.append(f"- `{s['file']}` · {s['name']}：{kv}")
    report.append("\n## ⚠️ 需人工确认")
    report.append("- **身份类型**：在校实习生等无法从「状态」列识别，请按实际核对；状态列缺失的仍默认「全日制正式」。")
    report.append("- **唯一标识**：若用工号，请确认同一人在不同主体的工号一致（不一致会把迁移的同一人拆成两人，建议改用身份证）。")
    if no_id:
        report.append(f"- **缺唯一标识(用姓名兜底，{len(no_id)}人)**：{('、'.join(no_id[:20]))}{'…' if len(no_id) > 20 else ''} —— 重名会错合，请补工号/身份证。")
    report.append("- **公积金现状**：原表无公积金列的，默认空（视为未缴）；如实际有缴存，请逐人补「是否缴/基数/比例」。")
    report.append("- **应发≠实发**：本表「应发工资」为税前合计；若原表只有实发，需补回个人社保/公积金/个税。")
    if need_manual:
        report.append("\n## 🤖 需 AI 人工接管的文件（脚本未能完全自动）")
        report.append("用 `inspect_payroll` 看这些文件的表头与样本，判断每列含义后，调用 `normalize_payroll` 时在 overrides 里传 column_map/entity/ym：")
        for fn, why in need_manual:
            report.append(f"- `{fn}` —— {why}")
    out_md = os.path.join(output_dir, '整理报告.md')
    with open(out_md, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report))

    return {
        'template_path': out_xlsx,
        'report_path': out_md,
        'records': len(records),
        'files_processed': len(files),
        'gross_methods': dict(gross_methods),
        'need_manual': [{'file': fn, 'reason': why} for fn, why in need_manual],
        'no_id_count': len(no_id),
        'stype_source': dict(stype_source),
        'unknown_status': {val: sorted(names) for val, names in sorted(unknown_status.items())},
        'ss_gap_never': [{'name': nm, 'entity': ent, 'months': v['zero']}
                         for (nm, ent), v in sorted(never_ss.items())],
        'ss_gap_partial': [{'name': nm, 'entity': ent, 'missing_months': v['zero'], 'paid_months': v['paid']}
                           for (nm, ent), v in sorted(partial_ss.items())],
        'suspect_rows': suspect_rows,
        'report_markdown': '\n'.join(report),
    }


def generate_blank_template(output_path=None):
    """生成空白标准模板（含顶部说明 + 身份类型/是否缴公积金下拉 + 示例行）。"""
    output_path = output_path or os.path.join(os.getcwd(), '社保测算-标准工资模板.xlsx')
    KIND = [1, 1, 1, 1, 1, 0, 2, 2, 2, 2]
    WIDTH = [16, 10, 14, 12, 18, 14, 14, 14, 16, 16]
    EXPLAIN = [
        ('社保测算 · 标准工资模板', True),
        ('① 一行 = 某人某月某主体的工资。同一人跨主体/跨月请分多行填写，系统会按"自然人"自动归并算月均。', False),
        ('② 加粗列为必填：工号/唯一标识、姓名、所属主体、所属年月(YYYY-MM)、应发工资(税前合计)。', False),
        ('③ 应发工资 = 税前合计(基本+岗位+绩效+奖金+补贴…)，不含单位社保；个人社保/个人公积金仍计入应发。切勿填实发。', False),
        ('④ 身份类型、是否缴公积金为下拉选择。身份类型留空默认"全日制正式"，可在工具第3步再调整。', False),
        ('⑤ 公积金三列仅在"现状已缴公积金"时填，用于诊断现状是否足额；不缴则留空或选"否"。', False),
        ('⑥ 删除下面示例行后再填写自己的数据。', False),
    ]
    HEADER_ROW = len(EXPLAIN) + 2
    EXAMPLE_ROW = HEADER_ROW + 1
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '标准模板'
    fill_req = PatternFill('solid', fgColor='3A5878'); fill_rec = PatternFill('solid', fgColor='E7EDF4')
    fill_opt = PatternFill('solid', fgColor='F1EEE7'); fill_band = PatternFill('solid', fgColor='FBFAF7')
    thin = Side(style='thin', color='E5E2DA'); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncol = len(COLS)
    for i, (text, is_title) in enumerate(EXPLAIN):
        r = i + 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(r, 1, text)
        c.font = Font(name='微软雅黑', size=15 if is_title else 10.5, bold=is_title, color='1B1D21' if is_title else '52545A')
        ws.row_dimensions[r].height = 28 if is_title else 20
        c.alignment = Alignment(horizontal='left', vertical='center'); c.fill = fill_band
    for j, (name, kind) in enumerate(zip(COLS, KIND), start=1):
        c = ws.cell(HEADER_ROW, j, name); c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if kind == 1:
            c.fill = fill_req; c.font = Font(name='微软雅黑', size=10.5, bold=True, color='FFFFFF')
        elif kind == 0:
            c.fill = fill_rec; c.font = Font(name='微软雅黑', size=10.5, bold=True, color='3A5878')
        else:
            c.fill = fill_opt; c.font = Font(name='微软雅黑', size=10.5, color='7C7E83')
        ws.column_dimensions[get_column_letter(j)].width = WIDTH[j - 1]
    ws.row_dimensions[HEADER_ROW].height = 30
    for j, v in enumerate(['1001', '张三', '示例科技有限公司', '2025-01', 12000, '全日制正式', '', '否', '', ''], start=1):
        c = ws.cell(EXAMPLE_ROW, j, v); c.border = border
        c.font = Font(name='微软雅黑', size=10.5, italic=True, color='9A9C9F')
        c.alignment = Alignment(horizontal='center', vertical='center')
    last = HEADER_ROW + 1000
    dv_type = DataValidation(type='list', formula1='"%s"' % ','.join(STYPES), allow_blank=True)
    ws.add_data_validation(dv_type); dv_type.add(f'F{EXAMPLE_ROW}:F{last}')
    dv_gjj = DataValidation(type='list', formula1='"是,否"', allow_blank=True)
    ws.add_data_validation(dv_gjj); dv_gjj.add(f'H{EXAMPLE_ROW}:H{last}')
    ws.freeze_panes = ws.cell(EXAMPLE_ROW, 1)
    wb.save(output_path)
    return output_path
